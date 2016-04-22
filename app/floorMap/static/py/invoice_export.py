# coding: utf-8

import os.path
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter

from django.conf import settings

import app.floorMap
from app.floorMap.models import Settings

APP_ROOT = app.floorMap.__path__[0]

TEMPLATE_FILE = "invoice_template.xlsx"
TEMPLATE_PATH = os.path.join("static", "xlsx", TEMPLATE_FILE)

WORKSHEETS = ['Som']

COLUMNS = [
    'name',
    'in_charge',
    'payment_method',
    'fees',
    'tps',
    'tvq',
    'total_fees',
    'phone',
    'email',
    'birthday',
    'rent_start',
    'rent_end',
]

HEADER_ROW = 4
SUM_ROW_DISTANCE = 1  # Number of rows between data and the sum row


def export_to_excel(invoice_rents, file_name):
    """
    Main method. Fills in a template excel file and saves to disk space
    :param invoice_rents: The rents to be exported
    :param invoice_unique_id: A unique ID to add to the exported filename
    :return: The name of the file
    """
    prev_company = None
    current_row = HEADER_ROW

    wb = load_template()
    ws = wb[WORKSHEETS[0]]

    for rent in invoice_rents:
        insert_new_row = rent.company != prev_company

        current_row += (1 if insert_new_row else 0)  # Bump row if needed

        # Get the row data
        row_data = get_row_data(ws, current_row)

        # Determine data to insert
        if insert_new_row:
            # Copy styles & pushes total
            insert_row(ws, row_data)

            tps = Settings.load().taxes_tps / 100
            tvq = Settings.load().taxes_tvq / 100

            # Data to insert
            data = dict(
                name=rent.company.name,
                in_charge="{first_name} {last_name}".format(
                    first_name=rent.company.person_in_charge.user.first_name,
                    last_name=rent.company.person_in_charge.user.last_name
                ),
                payment_method=rent.company.payment_method,
                fees="=" + str(rent.get_monthly_fee()),
                tps="=ROUND(D{row}*{tps}, 2)".format(row=current_row, tps=tps),
                tvq="=ROUND(D{row}*{tvq}, 2)".format(row=current_row, tvq=tvq),
                total_fees="=SUM(D{row}:F{row})".format(row=current_row),
                phone=rent.company.phone,
                email=rent.company.person_in_charge.user.email,
                birthday="?",
                # rent_start=rent.date_start.strftime('%b %Y'),
                # rent_end=rent.date_end.strftime('%b %Y'),
                rent_start="???",
                rent_end="???",
            )

            prev_company = rent.company
        else:
            # Data to insert
            data = dict(
                fees=" + ".join((
                    row_data[3].value,
                    str(rent.get_monthly_fee())
                ))
            )

        # Insert data into the row
        data_insert(row_data, data)

    # Saves file and exit
    return save_to_file(wb, file_name)


def load_template():
    """
    Load the template file
    :return: The loaded workbook
    """
    template_file = os.path.join(APP_ROOT, TEMPLATE_PATH)
    return load_workbook(filename=template_file)


def save_to_file(wb, file_name):
    """
    Save the workbook using the template filename added with an ID
    :param wb: The workbook to save
    :param file_name: File name
    :return: The export file path
    """
    export_path = os.path.join(settings.MEDIA_ROOT, file_name)
    wb.save(export_path)

    return export_path


def insert_row(ws, target_row):
    """
    Insert new row to the data table and pushes sum area
    :param ws: Current worksheet
    :param target_row: Target row number
    :return:
    """
    row_number = target_row[0].row

    if row_number != HEADER_ROW + 1:  # Do insert bump
        template_row = get_row_data(ws, HEADER_ROW + 1)
        sum_row = get_row_data(ws, row_number + SUM_ROW_DISTANCE + 1)
        old_sum_row = get_row_data(ws, row_number + SUM_ROW_DISTANCE)

        # Moves sum row styles
        copy_row(old_sum_row, sum_row)
        clear_row(old_sum_row)

        # Applies template row styles
        copy_row(template_row, target_row, styles_only=True)
    else:  # First insert, no bump
        sum_row = get_row_data(ws, row_number + SUM_ROW_DISTANCE + 1)

    # Sum data to insert
    data = dict(
        fees="=SUM(D{row_s}:D{row_e})".format(
            row_s=row_number, row_e=HEADER_ROW + 1),
        tps="=SUM(E{row_s}:E{row_e})".format(
            row_s=row_number, row_e=HEADER_ROW + 1),
        tvq="=SUM(F{row_s}:F{row_e})".format(
            row_s=row_number, row_e=HEADER_ROW + 1),
        total_fees="=SUM(G{row_s}:G{row_e})".format(
            row_s=row_number, row_e=HEADER_ROW + 1),
    )
    data_insert(sum_row, data)


def copy_row(source_row, target_row, data_only=False, styles_only=False):
    """
    Copies data and styles from one row to another
    :param source_row: The source row
    :param target_row: The target row
    :param data_only: Only copy cell value
    :param styles_only: Only copy cell styles
    :return:
    """
    for source_cell in source_row:
        ind = column_index_from_string(source_cell.column) - 1
        target_cell = target_row[ind]

        if not data_only:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format

        if not styles_only:
            target_cell.value = source_cell.value


def clear_row(target_row):
    """
    Erases styles and value on a row
    :param target_row: The where to erase styles
    :return:
    """
    for cell in target_row:
        cell.font = Font()
        cell.fill = PatternFill()
        cell.border = Border()
        cell.alignment = Alignment()
        cell.number_format = 'General'
        cell.value = ''


def data_insert(target_row, data):
    """
    Insert data into row
    :param target_row: Row where data is to be inserted
    :param data: The data to insert
    :return: The filled row
    """
    for key, value in data.items():
        column_id = COLUMNS.index(key)
        target_row[column_id].value = value


def get_row_data(ws, row):
    """
    Get the data cells on a specified row
    :param ws: The current worksheet
    :param row: The row containing the data
    :return: Tuple containing the row's data cells
    """
    coord_start = string_from_coordinate('A', row)
    coord_end = string_from_coordinate(len(COLUMNS), row)
    return ws[coord_start:coord_end][0]


def string_from_coordinate(column, row):
    """
    Reverse of utils function 'coordinate_from_string'
    :param column: column number
    :param row: row number
    :return: string coordinate
    """
    if isinstance(column, int):
        column = get_column_letter(column)
    return "{column}{row}".format(column=column, row=row)
